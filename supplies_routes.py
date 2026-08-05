from fastapi import APIRouter, Depends, HTTPException
from dependencies import verify_token, get_session
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from models import Product, User, StockMovement
from schemas import SuppliesSchema, ProductAdjustmentSchema, ProductUpdateSchema, StockMovementSchema, StockMovementResponse
from io import BytesIO
import pandas as pd
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

supplies_router = APIRouter(prefix='/supplies', tags=['supplies'], dependencies=[Depends(verify_token)])

# =========================
# HEALTH / HOME
# =========================

@supplies_router.get('/')
async def home():
    return {
        "message": "Rota inicial dos insumos"
    }

# =========================
# PRODUTOS - CONSULTAS
# =========================

@supplies_router.post('/products')
async def create_product(supplies_schema: SuppliesSchema, session: Session = Depends(get_session), user: User = Depends(verify_token)):

    """
        ## Criar produto (Admin)
        POST /products
    """
    product = session.query(Product).filter(Product.code == supplies_schema.code).first()
    if product:
        raise HTTPException(status_code=400, detail='O codigo do produto já existe')
    if user.role != "admin":
        raise HTTPException(status_code=401, detail='Voce nao tem autorizacao para realizar essa operacao.')
    new_product =  Product(supplies_schema.code, supplies_schema.product_type, supplies_schema.stock, supplies_schema.stock_minimum)
    session.add(new_product)
    session.commit()
    return {
        "message": f"Produto {new_product.code} adicionado com sucesso ",
        "product": new_product
    }

def build_products_query(
        session: Session,
        type: str | None = None,
        active: bool | None = None,
        search: str | None = None
):
    query = session.query(Product)

    if type:
        query = query.filter(func.lower(Product.product_type) == type.strip().lower())
    if active is not None:
        query = query.filter(Product.active == active)
    if search:
        query = query.filter(Product.code.ilike(f"%{search.strip()}%"))
    return query


@supplies_router.get('/products')
async def get_products(
    type: str | None = None, 
    active: bool | None = None, 
    search: str | None = None, 
    session: Session = Depends(get_session)
):
    query = build_products_query(
        session,
        type,
        active,
        search
    )
    return query.all()

@supplies_router.get('/products/export')
async def export_products(
    type: str | None = None, 
    active: bool | None = None, 
    search: str | None = None, 
    session: Session = Depends(get_session)
):
    products = build_products_query(
        session,
        type,
        active,
        search
    ).all()

    data = []
    for product in products:
        data.append({
        "ID": product.id,
        "Código": product.code,
        "Tipo": product.product_type,
        "Estoque": product.stock,
        "Estoque mínimo": product.stock_minimum,
        "Ativo": "Sim" if product.active else "Não",
        "Situação Estoque": (
            "Baixo estoque"
            if product.stock <= product.stock_minimum
            else "Normal"
        ),  
        "Criado em": product.created_at,
        "Atualizado em": product.updated_at
        })
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(
        output,
        index=False,
        engine="openpyxl"
    )
    output.seek(0)

    workbook = load_workbook(output)
    worksheet = workbook.active
    for cell in worksheet[1]:
        cell.font = Font(
            bold=True
        )
        cell.alignment = Alignment(
            horizontal="center"
        )
    for column in worksheet.columns:
        max_length = 0

        column_letter = get_column_letter(
            column[0].column
        )

        for cell in column:
            if cell.value:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        worksheet.column_dimensions[column_letter].width = max_length + 3

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"
    for row in worksheet.iter_rows(min_row=2):
        estoque = row[3].value
        minimo = row[4].value

        if estoque <= minimo:
            for cell in row:
                cell.fill = PatternFill(
                    "solid",
                    fgColor="FFC7CE"
                )
    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=produtos.xlsx"
        }
    )

@supplies_router.get('/products/low-stock')
async def low_stock(session: Session = Depends(get_session)):
    products = (
        session.query(Product).filter(Product.stock <= Product.stock_minimum).all()
    )
    if not products:
        return{
            "message": "Nenhum produto com estoque baixo",
            "products": []
        }
    return {
        "total": len(products),
        "products": products
    }


# =========================
# PRODUTO ESPECÍFICO
# =========================

@supplies_router.get('/products/{code}')
async def search_product(code: str, session: Session = Depends(get_session)):
    """
        ## Buscar um produto
        GET /products/{id}  -- busca um insumo a partir do ID
    """
    product = session.query(Product).filter(Product.code == code).first()
    if not product:
        raise HTTPException(status_code=404, detail='Codigo da etiqueta nao encontrada')
    return{
        "product": product
    }

@supplies_router.put('/products/{code}')
async def edit_product(code: str, product_update: ProductUpdateSchema ,session: Session = Depends(get_session)):
    """
        ## Editar um produto
        PUT /products/{id}
    """
    product = session.query(Product).filter(Product.code == code.upper()).first()
    if not product:
        raise HTTPException(status_code=404, detail='Produto nao encontrado')
    
    code_exists = (
    session.query(Product)
    .filter(Product.code == product_update.code)
    .filter(Product.id != product.id)
    .first()
)
    if code_exists:
        raise HTTPException(status_code=400, detail='O codigo já está cadastrado')
    
    product.code = product_update.code
    product.product_type = product_update.product_type
    session.commit()
    session.refresh(product)

    return {
        "message": "Produto atualizado com sucesso",
        "code": product.code, 
        "product_type": product.product_type, 
    }

@supplies_router.patch('/products/{code}/inactive')
async def inactive_product(code:str, session:Session = Depends(get_session)):
    product = session.query(Product).filter(Product.code == code).first()
    if not product:
        raise HTTPException(status_code=404, detail='Produto nao encontrado')
    product.active=False
    session.commit()
    session.refresh(product)
    return{
        "message": f"Produto {product.code} Desativado"
    }

@supplies_router.patch('/products/{code}/active')
async def active_product(code:str, session:Session = Depends(get_session)):
    product = session.query(Product).filter(Product.code == code).first()
    if not product:
        raise HTTPException(status_code=404, detail='Produto nao encontrado')
    product.active=True
    session.commit()
    session.refresh(product)
    return{
        "message": f"Produto {product.code} Ativado"
    }

# =========================
# MOVIMENTAÇÕES
# =========================

@supplies_router.post('/products/{code}/entry')
async def entry_product(code: str, movement_schema: StockMovementSchema,session: Session = Depends(get_session), user: User = Depends(verify_token)):
    if user.role != "admin":
        raise HTTPException(status_code=401, detail='Voce nao tem autorizacao para realizar essa operacao.')
    product = session.query(Product).filter(Product.code == code).first()
    if not product:
        raise HTTPException(status_code=404, detail='Produto nao encontrado')
    stock_before = product.stock
    product.stock += movement_schema.quantity
    movement = StockMovement(
        product_id = product.id,
        user_id = user.id,
        movement_type = "entrada",
        quantity = movement_schema.quantity,
        stock_before = stock_before,
        stock_after = product.stock,
        observation = movement_schema.observation
    )

    session.add(movement)
    session.commit()
    session.refresh(product)
    return {
        "message": "Entrada registrada com sucesso",
        "stock": product.stock
    }

@supplies_router.post('/products/{code}/exit')
async def exit_product(code: str, movement_schema: StockMovementSchema, session: Session = Depends(get_session), user: User = Depends(verify_token)):
    if user.role != "admin":
        raise HTTPException(status_code=401, detail='Voce nao tem autorizacao para realizar essa operacao.')
    product = session.query(Product).filter(Product.code == code).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produto nao encontrado")
    if product.stock < movement_schema.quantity:
        raise HTTPException(status_code=404, detail="Estoque insuficiente")
    stock_before = product.stock
    product.stock -= movement_schema.quantity
    movement = StockMovement(
        product_id = product.id,
        user_id = user.id,
        movement_type = "saida",
        quantity = movement_schema.quantity,
        stock_before = stock_before,
        stock_after = product.stock,
        observation = movement_schema.observation
    )
    session.add(movement)
    session.commit()
    session.refresh(product)
    return {
        "message": "Saida registrada com sucesso",
        "stock": product.stock
    }

@supplies_router.get('/products/{code}/movements')
async def product_movements(
    code: str,
    session: Session = Depends(get_session)
):
    product = (
        session.query(Product)
        .filter(Product.code == code)
        .first()
    )

    if not product:
        raise HTTPException(
            status_code=404,
            detail="Produto nao encontrado"
        )

    movements = (
        session.query(StockMovement)
        .options(joinedload(StockMovement.user))
        .filter(StockMovement.product_id == product.id)
        .order_by(StockMovement.created_at.desc())
        .all()
    )

    return {
        "product": {
            "code": product.code,
            "current_stock": product.stock
        },
        "movements": [StockMovementResponse.from_movement(m) for m in movements]
    }

# =========================
# AJUSTE
# =========================

@supplies_router.post('/products/{id}/adjustment')
async def adjustment(
    id: int,
    product_adjustment: ProductAdjustmentSchema,
    session: Session = Depends(get_session),
    user: User = Depends(verify_token)
):
    if user.role != "admin":
        raise HTTPException(status_code=401, detail='Voce nao tem autorizacao para realizar essa operacao.')
    product = session.query(Product).filter(Product.id == id).first()
    if not product:
        raise HTTPException(404, "Produto nao encontrado")
    if product_adjustment.new_stock < 0:
        raise HTTPException(400, "O estoque nao pode ser negativo")
    if product_adjustment.new_stock == product.stock:
        raise HTTPException(400, "O estoque informado eh igual ao estoque atual.")

    difference = product_adjustment.new_stock - product.stock
    movement = StockMovement(
        product_id=product.id,
        user_id=user.id,
        type="adjustment",
        quantity=difference,
        reason=product_adjustment.reason
    )
    product.stock = product_adjustment.new_stock

    session.add(movement)
    session.commit()
    session.refresh(product)

    return {
        "message": "Estoque ajustado com sucesso.",
        "difference": difference,
        "current_stock": product.stock
    }